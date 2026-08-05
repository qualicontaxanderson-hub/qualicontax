# -*- coding: utf-8 -*-
"""
Gerador de RELATÓRIO (Entradas / Saídas / CT-e) — PDF (ReportLab) e Excel (openpyxl).
Python puro. Recebe dados já prontos (o app monta a partir do filtro) e devolve bytes.

API:
    from utils.relatorio import gerar_relatorio_pdf, gerar_relatorio_xlsx
    ctx = {
      'titulo': 'Relatório de Saídas',
      'empresa': '162 - POSTO NOVO HORIZONTE GOIATUBA LTDA',
      'periodo': '01/08/2026 a 05/08/2026',
      'gerado_em': '05/08/2026 10:45',
      'kpis': [('Total de Notas','9.334'), ('Valor Total','R$ 3.810.355,34'), ...],  # até 5
      'colunas': [('Data / Hora','L'), ('Nº/Série','L'), ('Valor R$','R'), ...],      # L=esq R=dir
      'linhas': [ ['05/08/2026 10:31','1481/13', ... ], ... ],   # strings já formatadas
      'totais': ['','','TOTAL', ..., 'R$ ...'],                  # mesma largura de colunas
      'logo_path': '/caminho/logo.png',                          # opcional
      'aviso': None,   # ou 'Listagem truncada em 5.000 linhas — use o Excel para o total.'
    }
    pdf_bytes  = gerar_relatorio_pdf(ctx)
    xlsx_bytes = gerar_relatorio_xlsx(ctx)   # colunas 'R' saem numéricas se vier número
"""
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Image,
                                Paragraph, Spacer)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

GREEN = colors.HexColor('#16A34A')
DGREEN = colors.HexColor('#15803D')
ZEBRA = colors.HexColor('#F8FAFC')
TOTBG = colors.HexColor('#DCFCE7')
GREY = colors.HexColor('#64748B')
LINE = colors.HexColor('#EEF2F7')
KPI_CORES = [colors.HexColor('#334155'), GREEN, colors.HexColor('#2563EB'),
             colors.HexColor('#D97706'), colors.HexColor('#DC2626')]

def _p(txt, size=8, color=colors.black, bold=False, align=TA_LEFT):
    st = ParagraphStyle('x', fontName='Helvetica-Bold' if bold else 'Helvetica',
                        fontSize=size, leading=size + 2, textColor=color, alignment=align)
    return Paragraph(str(txt), st)

def gerar_relatorio_pdf(ctx):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=8 * mm, bottomMargin=12 * mm,
                            leftMargin=8 * mm, rightMargin=8 * mm,
                            title=ctx.get('titulo', 'Relatório'))
    larg = doc.width
    el = []

    # ---- cabeçalho: logo | título | meta ----
    logo_cell = ''
    if ctx.get('logo_path'):
        try:
            from reportlab.lib.utils import ImageReader
            iw, ih = ImageReader(ctx['logo_path']).getSize()
            h = 16 * mm; w = h * iw / ih
            logo_cell = Image(ctx['logo_path'], width=w, height=h)
        except Exception:
            logo_cell = ''
    titulo = [_p(ctx.get('titulo', 'Relatório'), 17, DGREEN, bold=True),
              _p('Qualicontax — Assessoria Contábil', 9, GREY)]
    meta = _p('Gerado em %s' % ctx.get('gerado_em', ''), 8, GREY, align=TA_RIGHT)
    head = Table([[logo_cell, titulo, meta]], colWidths=[40 * mm, larg - 100 * mm, 60 * mm])
    head.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 2, GREEN),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    el += [head, Spacer(1, 6)]

    # ---- info empresa/período ----
    el += [_p('<b>Empresa:</b> %s&nbsp;&nbsp;&nbsp;&nbsp;<b>Período:</b> %s'
              % (ctx.get('empresa', '—'), ctx.get('periodo', '—')), 9), Spacer(1, 6)]

    # ---- KPIs coloridos ----
    kpis = ctx.get('kpis', [])[:5]
    if kpis:
        cells = [[_p(l.upper(), 7, colors.white), _p(v, 13, colors.white, bold=True)]
                 for l, v in kpis]
        inner = [Table([[c[0]], [c[1]]]) for c in cells]
        krow = Table([inner], colWidths=[(larg) / len(inner)] * len(inner))
        ts = [('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
              ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 4)]
        krow.setStyle(TableStyle(ts))
        for i in range(len(inner)):
            inner[i].setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), KPI_CORES[i % len(KPI_CORES)]),
                ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('ROUNDEDCORNERS', [4, 4, 4, 4]),
            ]))
        el += [krow, Spacer(1, 10)]

    # ---- tabela ----
    cols = ctx.get('colunas', [])
    aligns = [TA_RIGHT if c[1] == 'R' else TA_LEFT for c in cols]
    header = [_p(c[0], 7.5, colors.white, bold=True, align=aligns[i]) for i, c in enumerate(cols)]
    body = [header]
    for row in ctx.get('linhas', []):
        body.append([_p(row[i], 8, align=aligns[i]) for i in range(len(cols))])
    totais = ctx.get('totais')
    if totais:
        body.append([_p(totais[i], 8, bold=True, align=aligns[i]) for i in range(len(cols))])

    tbl = Table(body, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), DGREEN),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 1), (-1, -2), 0.4, LINE),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    for r in range(1, len(body) - (1 if totais else 0)):
        if r % 2 == 0:
            style.append(('BACKGROUND', (0, r), (-1, r), ZEBRA))
    if totais:
        style += [('BACKGROUND', (0, -1), (-1, -1), TOTBG),
                  ('LINEABOVE', (0, -1), (-1, -1), 1.2, GREEN)]
    tbl.setStyle(TableStyle(style))
    el += [tbl]

    if ctx.get('aviso'):
        el += [Spacer(1, 8), _p('⚠ ' + ctx['aviso'], 8, colors.HexColor('#B45309'))]

    def _rodape(canvas, d):
        canvas.saveState()
        canvas.setFont('Helvetica', 7); canvas.setFillColor(GREY)
        canvas.drawString(8 * mm, 6 * mm, 'Qualicontax Assessoria Contábil — Relatório gerado pelo sistema')
        canvas.drawRightString(d.pagesize[0] - 8 * mm, 6 * mm, 'Página %d' % canvas.getPageNumber())
        canvas.restoreState()

    doc.build(el, onFirstPage=_rodape, onLaterPages=_rodape)
    return buf.getvalue()


def gerar_relatorio_xlsx(ctx):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        XLImage = None

    wb = Workbook(); ws = wb.active; ws.title = 'Dados'
    cols = ctx.get('colunas', [])
    hdr_fill = PatternFill('solid', fgColor='15803D')
    white_bold = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    tot_fill = PatternFill('solid', fgColor='DCFCE7')
    border = Border(bottom=Side(style='thin', color='E2E8F0'))

    if XLImage and ctx.get('logo_path'):
        try:
            im = XLImage(ctx['logo_path']); im.height = 60; im.width = 100
            ws.add_image(im, 'A1'); ws.row_dimensions[1].height = 48
        except Exception:
            pass
    ws['C1'] = ctx.get('titulo', 'Relatório'); ws['C1'].font = Font(name='Arial', bold=True, size=15, color='15803D')
    ws['C2'] = 'Qualicontax — Assessoria Contábil'; ws['C2'].font = Font(name='Arial', size=10, color='64748B')
    ws['C4'] = 'Empresa:'; ws['C4'].font = Font(name='Arial', bold=True, size=10)
    ws['D4'] = ctx.get('empresa', ''); ws['D4'].font = Font(name='Arial', size=10)
    ws['C5'] = 'Período:'; ws['C5'].font = Font(name='Arial', bold=True, size=10)
    ws['D5'] = ctx.get('periodo', ''); ws['D5'].font = Font(name='Arial', size=10)

    hr = 7
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=hr, column=i, value=c[0])
        cell.font = white_bold; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='right' if c[1] == 'R' else 'left', vertical='center')

    r = hr + 1
    for row in ctx.get('linhas', []):
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = Font(name='Arial', size=10); cell.border = border
            if cols[i - 1][1] == 'R':
                cell.alignment = Alignment(horizontal='right')
                if isinstance(v, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
        r += 1

    # totais por SOMA (fórmula) nas colunas numéricas
    last = r - 1
    ws.cell(row=r, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
    for i, c in enumerate(cols, 1):
        if c[1] == 'R' and last >= hr + 1:
            L = get_column_letter(i)
            cell = ws.cell(row=r, column=i, value=f'=SUM({L}{hr+1}:{L}{last})')
            cell.font = Font(name='Arial', bold=True); cell.number_format = 'R$ #,##0.00'
            cell.alignment = Alignment(horizontal='right')
    for i in range(1, len(cols) + 1):
        ws.cell(row=r, column=i).fill = tot_fill

    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 34 if i == 3 else 14
    ws.freeze_panes = f'A{hr+1}'
    out = io.BytesIO(); wb.save(out); return out.getvalue()
