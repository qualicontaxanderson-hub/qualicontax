"""Rotas do módulo Configurações — Usuários e Perfis de Acesso"""
import logging
import re
import threading
from io import BytesIO
from datetime import date, datetime, timedelta, timezone

from flask import (Blueprint, render_template, request, redirect, url_for, flash,
                   jsonify, send_file)
from flask_login import current_user
from utils.auth_helper import login_required, admin_required, permission_required, hash_password
from utils.atividade import registrar
from utils.auditoria_fmt import (hist_preparar, coletar_cliente_ids,
                                 MODULO_LABEL, AUDITORIA_INICIO)
from utils.db_helper import execute_query, transacao
from utils.permissions import PERMISSION_CATALOG
from utils import cadastro_token
from utils import colabore_chaves

logger = logging.getLogger(__name__)

configuracoes = Blueprint('configuracoes', __name__, url_prefix='/configuracoes')


# ---------------------------------------------------------------------------
# Index
# ---------------------------------------------------------------------------
@configuracoes.route('/')
@permission_required('configuracoes.index')
def index():
    # Espaço do Dropbox: SÓ para admin e SÓ leitura (users/get_space_usage), com
    # cache de 1h em app_config — o render normal não bate na API do Dropbox.
    # Consultado apenas para admin para que um refresh de não-admin nem chegue
    # perto de gastar rate limit.
    espaco = None
    if current_user.is_admin():
        from utils.dropbox_space import get_space
        espaco = get_space()
    return render_template('configuracoes/index.html', dropbox_espaco=espaco)


# ---------------------------------------------------------------------------
# Home do Configurações — carrossel e contadores.
#
# Mesmo contrato JSON das homes de Cadastros e Escrita Fiscal (cards /
# counters / gerado_em_ms), para reaproveitar static/js/home_destaques.js sem
# copiar uma linha de JavaScript.
#
# Regra da casa: NADA de número inventado. Card cuja fonte não responde
# simplesmente não entra — e nenhum card ganha sparkline sem série real por
# trás. Aqui isso pesa mais que nas outras telas, porque os dois indicadores
# de armazenamento existem para DECIDIR (pagar mais, limpar, migrar): um
# número errado sai caro.
# ---------------------------------------------------------------------------
_CFG_HOME_CACHE: dict = {}
_CFG_HOME_CACHE_LOCK = threading.Lock()
_CFG_HOME_TTL_S = 60


def _cfg_q1(sql, params=None):
    """Escalar isolado: query que falha derruba o card dela, não a tela."""
    try:
        return execute_query(sql, tuple(params or ()), fetch=True, fetch_one=True) or {}
    except Exception:
        logger.exception('[config-home] query omitida')
        return {}


def _cfg_i(d, k):
    v = (d or {}).get(k)
    return int(v) if v is not None else 0


def _cfg_trend(atual, anterior):
    if not anterior:
        return {'tipo': 'neutro', 'rotulo': 'no mês'}
    pct = round((atual - anterior) * 100.0 / anterior)
    if pct > 0:
        return {'tipo': 'alta', 'pct': pct}
    if pct < 0:
        return {'tipo': 'baixa', 'pct': pct}
    return {'tipo': 'igual', 'pct': 0}


def _config_home_payload():
    from utils.dropbox_space import formatar_bytes

    cards = []

    # ---- Usuários e perfis -------------------------------------------------
    u = _cfg_q1(
        "SELECT "
        " (SELECT COUNT(*) FROM usuarios)                                  total,"
        " (SELECT COUNT(*) FROM usuarios WHERE situacao='ATIVO')           ativos,"
        " (SELECT COUNT(*) FROM usuarios WHERE tipo_usuario='ADMIN')       admins,"
        " (SELECT COUNT(*) FROM usuarios WHERE senha_pendente=1)           pendentes,"
        " (SELECT COUNT(*) FROM perfis_acesso)                             perfis,"
        " (SELECT COUNT(*) FROM perfis_acesso WHERE situacao='ATIVO')      perfis_at")
    total_u, ativos_u = _cfg_i(u, 'total'), _cfg_i(u, 'ativos')
    admins, pendentes = _cfg_i(u, 'admins'), _cfg_i(u, 'pendentes')
    perfis, perfis_at = _cfg_i(u, 'perfis'), _cfg_i(u, 'perfis_at')

    # Série real de usuários criados por mês (usuarios.criado_em).
    serie = execute_query(
        "SELECT DATE_FORMAT(criado_em,'%Y-%m') ym, COUNT(*) n FROM usuarios "
        "WHERE criado_em >= DATE_FORMAT(CURDATE() - INTERVAL 5 MONTH,'%Y-%m-01') "
        "GROUP BY ym ORDER BY ym", fetch=True) or []
    por_mes = {r['ym']: int(r['n']) for r in serie}
    hoje = date.today()
    meses = []
    for i in range(5, -1, -1):
        m = hoje.month - i
        a = hoje.year + (m - 1) // 12
        m = (m - 1) % 12 + 1
        meses.append('%04d-%02d' % (a, m))
    spark_u = [por_mes.get(m, 0) for m in meses]

    apoio_u = '%d ativo%s · %d administrador%s' % (
        ativos_u, '' if ativos_u == 1 else 's',
        admins, '' if admins == 1 else 'es')
    if pendentes:
        apoio_u += ' · %d com senha pendente' % pendentes
    cards.append({
        'titulo': 'Usuários cadastrados', 'icone': 'fa-users-cog',
        'valor': total_u, 'apoio': apoio_u,
        'trend': _cfg_trend(spark_u[-1], spark_u[-2]),
        # Só desenha a série se ela tiver algum movimento: uma linha reta em
        # zero é decoração se passando por informação.
        **({'spark': spark_u, 'spark_tipo': 'linha'} if any(spark_u) else {}),
    })

    cards.append({
        'titulo': 'Perfis de acesso', 'icone': 'fa-shield-halved',
        'valor': perfis,
        'apoio': '%d ativo%s · define o que cada um enxerga'
                 % (perfis_at, '' if perfis_at == 1 else 's'),
        'trend': {'tipo': 'neutro', 'rotulo': 'total'},
    })

    # ---- Armazenamento: Dropbox -------------------------------------------
    # Só entra quando a consulta deu certo. Com credencial inválida o card
    # SOME — o motivo já aparece em destaque no card detalhado da tela, e um
    # "0 GB" aqui pareceria disco vazio em vez de leitura indisponível.
    try:
        from utils.dropbox_space import get_space
        esp = get_space()
        if esp and esp.get('ok') and esp.get('usado') is not None:
            usado, total_b = int(esp['usado']), int(esp.get('total') or 0)
            livre = max(total_b - usado, 0)
            cards.append({
                'titulo': 'Dropbox usado', 'icone': 'fa-cloud',
                'valor': int(round(esp.get('pct') or 0)), 'valor_sufixo': '%',
                'apoio': '%s de %s · %s livre%s'
                         % (formatar_bytes(usado), formatar_bytes(total_b),
                            formatar_bytes(livre),
                            ' (valor anterior)' if esp.get('stale') else ''),
                'trend': {'tipo': 'neutro', 'rotulo': 'nuvem'},
                'spark': [usado, livre], 'spark_tipo': 'barra',
            })
    except Exception:
        logger.exception('[config-home] card do Dropbox omitido')

    # ---- Armazenamento: banco no Railway -----------------------------------
    # É o que cresce e o que se paga. NÃO se mostra percentual: o plano do
    # Railway não expõe cota por aqui, e inventar um teto para desenhar barra
    # seria número falso. Mostra-se o tamanho real e de onde ele vem.
    tam = _cfg_q1(
        "SELECT COUNT(*) tabelas, "
        "       CAST(COALESCE(SUM(data_length + index_length),0) AS UNSIGNED) bytes "
        "  FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE()")
    db_bytes, db_tabelas = _cfg_i(tam, 'bytes'), _cfg_i(tam, 'tabelas')
    maiores = execute_query(
        "SELECT TABLE_NAME t, "
        "       CAST((data_length + index_length) AS UNSIGNED) b "
        "  FROM information_schema.TABLES WHERE TABLE_SCHEMA = DATABASE() "
        " ORDER BY (data_length + index_length) DESC LIMIT 12", fetch=True) or []

    if db_bytes:
        # Valor em GB quando passa de 1 GB; abaixo disso, MB — número redondo
        # de ler, sem esconder a ordem de grandeza.
        if db_bytes >= 1024 ** 3:
            valor, sufixo = round(db_bytes / (1024.0 ** 3), 1), ' GB'
        else:
            valor, sufixo = int(round(db_bytes / (1024.0 ** 2))), ' MB'
        maior = maiores[0] if maiores else None
        cards.append({
            'titulo': 'Banco no Railway', 'icone': 'fa-database',
            'valor': valor, 'valor_sufixo': sufixo,
            'apoio': ('%d tabelas · maior: %s (%s)'
                      % (db_tabelas, maior['t'], formatar_bytes(int(maior['b'])))
                      if maior else '%d tabelas' % db_tabelas),
            'trend': {'tipo': 'neutro', 'rotulo': 'em disco'},
        })

    if maiores:
        cards.append({
            'titulo': 'Maiores tabelas', 'icone': 'fa-database', 'tipo': 'lista',
            'itens': [{'valor': formatar_bytes(int(r['b'])), 'rotulo': r['t'],
                       'barra': (int(r['b']) / float(maiores[0]['b']))
                                if int(maiores[0]['b']) else 0}
                      for r in maiores],
            'trend': {'tipo': 'neutro', 'rotulo': 'em disco'},
        })

    counters = {
        'usuarios': total_u,
        'perfis': perfis,
    }
    return {
        'cards': cards,
        'counters': counters,
        'gerado_em_ms': int(datetime.now(timezone.utc).timestamp() * 1000),
    }


@configuracoes.route('/api/destaques')
@permission_required('configuracoes.index')
def api_config_destaques():
    """Números da home do Configurações. Cache de 60s por usuário — o mesmo
    molde do Cadastros, e o que impede a consulta de tamanho do banco de rodar
    a cada refresh."""
    uid = getattr(current_user, 'id', None)
    agora = datetime.now(timezone.utc).timestamp()
    with _CFG_HOME_CACHE_LOCK:
        hit = _CFG_HOME_CACHE.get(uid)
        if hit and (agora - hit[0]) < _CFG_HOME_TTL_S:
            return jsonify(hit[1])
    payload = _config_home_payload()
    with _CFG_HOME_CACHE_LOCK:
        _CFG_HOME_CACHE[uid] = (agora, payload)
    return jsonify(payload)


# ---------------------------------------------------------------------------
# D4 — Auditoria (tela geral de movimentações). SÓ ADMIN (trava no servidor).
# ---------------------------------------------------------------------------
_AUD_MODULOS = ['fiscal', 'cadastros', 'contabil', 'dp', 'colabore', 'configuracoes']
_AUD_PER = 50
# Teto do export. Auditoria pode crescer; 20k é seguro no openpyxl e nunca corta
# calado (avisa na tela e na 1ª linha da planilha se o filtro tiver mais).
_AUD_EXPORT_CAP = 20000


def _aud_filtros():
    """Filtros da query string. Período padrão: últimos 7 dias."""
    hoje = date.today()
    return {
        'pessoa': request.args.get('pessoa', '').strip(),
        'modulo': request.args.get('modulo', '').strip(),
        'tipo': request.args.get('tipo', '').strip(),          # ''|escrita|leitura
        'empresa': request.args.get('empresa', '').strip(),
        # OPCIONAL, desmarcado por padrão: auditoria mostra TUDO por padrão (esconder
        # linha não audita). Só filtra teste quem marcar a caixa.
        'ocultar_teste': '1' if request.args.get('ocultar_teste') else '',
        'data_ini': request.args.get('data_ini', '').strip() or (hoje - timedelta(days=7)).isoformat(),
        'data_fim': request.args.get('data_fim', '').strip() or hoje.isoformat(),
    }


def _aud_where(f):
    """(cond, params) do WHERE. ZZ TESTE só sai se a pessoa PEDIR (ocultar_teste)."""
    cond = ["l.data_hora >= %s", "l.data_hora <= %s"]
    params = [f['data_ini'] + ' 00:00:00', f['data_fim'] + ' 23:59:59']
    if f['ocultar_teste']:
        cond.append("COALESCE(l.usuario_nome,'') NOT LIKE 'ZZ TESTE%%'")
    if f['pessoa']:
        cond.append("l.usuario_nome = %s"); params.append(f['pessoa'])
    if f['modulo']:
        cond.append("l.modulo = %s"); params.append(f['modulo'])
    if f['tipo'] == 'escrita':
        cond.append("l.acao LIKE 'escrita.%%'")
    elif f['tipo'] == 'leitura':
        cond.append("l.acao LIKE 'leitura.%%'")
    if f['empresa']:
        emp = execute_query(
            "SELECT id, numero_cliente FROM clientes "
            "WHERE numero_cliente = %s OR nome_razao_social LIKE %s",
            (f['empresa'], '%' + f['empresa'] + '%'), fetch=True) or []
        ids = [e['id'] for e in emp]
        nums = [str(e['numero_cliente']) for e in emp if e.get('numero_cliente')]
        if not ids and not nums:
            cond.append("1=0")                       # empresa não encontrada -> vazio
        else:
            ors, p2 = [], []
            if ids:
                ph = ','.join(['%s'] * len(ids))
                ors.append("(l.tabela_afetada IN ('clientes','dfe_certificados') "
                           f"AND l.registro_id IN ({ph}))")
                p2 += ids
            if nums:
                ph = ','.join(['%s'] * len(nums))
                for path in ('$.cliente_numero', '$.filtros.cliente_numero',
                             '$.numero_cliente', '$.numero'):
                    ors.append(f"JSON_UNQUOTE(JSON_EXTRACT(l.dados_novos,'{path}')) IN ({ph})")
                    p2 += nums
            cond.append('(' + ' OR '.join(ors) + ')')
            params += p2
    return ' AND '.join(cond), params


def _aud_resolver(rows):
    """{usuario_id: nome} para linhas antigas + {registro_id: '#num nome'} p/ empresa."""
    nomes = {}
    faltam = {r['usuario_id'] for r in rows if not r.get('usuario_nome') and r.get('usuario_id')}
    if faltam:
        ph = ','.join(['%s'] * len(faltam))
        for u in execute_query(f"SELECT id, nome FROM usuarios WHERE id IN ({ph})",
                               tuple(faltam), fetch=True) or []:
            nomes[u['id']] = u['nome']
    # Todos os cliente_id da página (registro_id + JSON: inclui filtros.cliente_id
    # das buscas antigas, que só têm o id cru). Resolvidos em UMA consulta.
    cids = coletar_cliente_ids(rows)
    emp_map = {}
    if cids:
        ph = ','.join(['%s'] * len(cids))
        for c in execute_query(
                f"SELECT id, numero_cliente, nome_razao_social FROM clientes WHERE id IN ({ph})",
                tuple(cids), fetch=True) or []:
            num, nome = c.get('numero_cliente'), (c.get('nome_razao_social') or '').strip()
            emp_map[c['id']] = ('#%s %s' % (num, nome)).strip() if num else nome
    return nomes, emp_map


_AUD_SELECT = ("SELECT l.id, l.data_hora, l.usuario_id, l.usuario_nome, l.acao, l.modulo, "
               "l.tabela_afetada, l.registro_id, l.dados_anteriores, l.dados_novos "
               "FROM logs_sistema l WHERE ")


def _aud_qs(f):
    from urllib.parse import urlencode
    return urlencode({k: v for k, v in f.items() if v})


@configuracoes.route('/auditoria')
@admin_required
def auditoria():
    """Tela geral de auditoria — todas as movimentações de logs_sistema."""
    f = _aud_filtros()
    cond, params = _aud_where(f)

    if request.args.get('export') == 'xlsx':
        return _aud_export(f, cond, params)

    try:
        page = max(1, int(request.args.get('page', 1) or 1))
    except (TypeError, ValueError):
        page = 1
    off = (page - 1) * _AUD_PER

    total = (execute_query(f"SELECT COUNT(*) AS c FROM logs_sistema l WHERE {cond}",
                           tuple(params), fetch=True, fetch_one=True) or {}).get('c', 0)
    rows = execute_query(_AUD_SELECT + cond + " ORDER BY l.data_hora DESC, l.id DESC LIMIT %s OFFSET %s",
                         tuple(params + [_AUD_PER, off]), fetch=True) or []
    nomes, emp_map = _aud_resolver(rows)
    itens = [hist_preparar(r, nomes, emp_map) for r in rows]

    # Resumo (participação por pessoa no filtro) — o mesmo número do card.
    rr = execute_query(f"SELECT COALESCE(l.usuario_nome,'—') AS nome, COUNT(*) AS n "
                       f"FROM logs_sistema l WHERE {cond} GROUP BY l.usuario_nome ORDER BY n DESC",
                       tuple(params), fetch=True) or []
    rtot = sum(int(r['n']) for r in rr)
    resumo = [{'nome': r['nome'], 'n': int(r['n']),
               'pct': round(100 * int(r['n']) / rtot) if rtot else 0} for r in rr]

    # Mostra TODOS os usuários com registro (inclusive teste) — a auditoria não
    # esconde ninguém; ocultar teste é opção da própria pessoa.
    pessoas = [r['usuario_nome'] for r in execute_query(
        "SELECT DISTINCT usuario_nome FROM logs_sistema WHERE usuario_nome IS NOT NULL "
        "ORDER BY usuario_nome", fetch=True) or []]

    total_paginas = max(1, (total + _AUD_PER - 1) // _AUD_PER)
    return render_template('configuracoes/auditoria.html',
                           itens=itens, resumo=resumo, resumo_total=rtot, total=total,
                           page=page, total_paginas=total_paginas, f=f, pessoas=pessoas,
                           modulos=_AUD_MODULOS, modulo_label=MODULO_LABEL,
                           auditoria_inicio=AUDITORIA_INICIO, qs=_aud_qs(f),
                           export_cap=_AUD_EXPORT_CAP)


def _aud_campos_txt(campos):
    parts = []
    for c in campos:
        if c['tipo'] == 'sensivel':
            parts.append('%s alterada' % c['label'])
        elif c['tipo'] == 'novo':
            parts.append('%s: %s' % (c['label'], c.get('depois')))
        elif c['tipo'] == 'removido':
            parts.append('%s: %s' % (c['label'], c.get('antes')))
        else:
            parts.append('%s: %s -> %s' % (c['label'], c.get('antes'), c.get('depois')))
    return ' | '.join(parts)


def _aud_export(f, cond, params):
    """XLSX do conjunto filtrado. NUNCA trunca calado: se o filtro tiver mais que
    o teto, avisa na 1ª linha da planilha (e a tela já avisou antes). O export
    TAMBÉM é participação: registra leitura."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    total = (execute_query(f"SELECT COUNT(*) AS c FROM logs_sistema l WHERE {cond}",
                           tuple(params), fetch=True, fetch_one=True) or {}).get('c', 0)
    rows = execute_query(_AUD_SELECT + cond + " ORDER BY l.data_hora DESC, l.id DESC LIMIT %s",
                         tuple(params + [_AUD_EXPORT_CAP]), fetch=True) or []
    nomes, emp_map = _aud_resolver(rows)
    itens = [hist_preparar(r, nomes, emp_map) for r in rows]
    truncou = total > _AUD_EXPORT_CAP

    wb = Workbook(); ws = wb.active; ws.title = 'Auditoria'
    cols = ['Data/hora', 'Usuário', 'Ação', 'Módulo', 'Empresa', 'Mudanças']
    if truncou:
        ws.append(['ATENÇÃO: o filtro tem %d registros; este arquivo traz os %d mais '
                   'recentes. Refine o período para levar tudo.' % (total, _AUD_EXPORT_CAP)])
        ws.cell(row=1, column=1).font = Font(bold=True, color='B45309')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.append(cols)
    hrow = ws.max_row
    for i in range(1, len(cols) + 1):
        cell = ws.cell(row=hrow, column=i)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='15803D')
    for it in itens:
        ws.append([it['data_hora'], it['autor'], it['verbo'],
                   MODULO_LABEL.get(it.get('modulo'), it.get('modulo') or ''),
                   it.get('empresa') or '', _aud_campos_txt(it['campos'])])
    for i, w in enumerate((17, 22, 28, 13, 34, 70), 1):
        ws.column_dimensions[chr(64 + i)].width = w
    bio = BytesIO(); wb.save(bio); bio.seek(0)

    registrar('leitura.exportou_arquivo', 'configuracoes', tabela='logs_sistema',
              depois={'formato': 'xlsx', 'escopo': 'auditoria', 'total_filtro': total,
                      'trazidos': len(itens), 'truncou': truncou,
                      'filtros': {k: v for k, v in f.items() if v}})
    return send_file(bio, as_attachment=True,
                     download_name='auditoria_%s_a_%s.xlsx' % (f['data_ini'], f['data_fim']),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------------------------------------------------------------------------
# F1 — Caixa de entrada (_ENTRADA do Dropbox). SÓ ADMIN. SÓ metadados: nunca
# baixa/abre arquivo, nunca grava o NOME no log (nomes de .pfx podem conter a
# senha do certificado). Só listar e renomear.
# ---------------------------------------------------------------------------
_ENTRADA_TIPOS = {
    'pfx': ('Certificado', 'fa-certificate'),
    'xml': ('Nota fiscal (XML)', 'fa-file-code'),
    'pdf': ('Documento (PDF)', 'fa-file-pdf'),
    'zip': ('Compactado (ZIP)', 'fa-file-zipper'),
}
_ENTRADA_SEP = ('-', '_', ' ', '.')


def _ent_sem_ext(nome):
    return nome.rsplit('.', 1)[0] if '.' in nome else nome


def _ent_casa_empresa(nome, empresas):
    """'#num NOME' da empresa cujo NÚMERO é prefixo (com separador) ou cujo
    CNPJ/CPF está contido (fronteira de dígitos) no nome. Mesma régua do
    certificado, generalizada p/ qualquer extensão. None se não casar."""
    base = _ent_sem_ext(nome)
    base_l = base.lower()
    for e in empresas:
        num = (e.get('numero_cliente') or '').strip()
        doc = re.sub(r'\D', '', e.get('cpf_cnpj') or '')
        casa_num = bool(num) and (base_l == num.lower() or (
            base_l.startswith(num.lower()) and len(base_l) > len(num)
            and base_l[len(num)] in _ENTRADA_SEP))
        casa_doc = len(doc) in (11, 14) and re.search(
            r'(?<!\d)' + re.escape(doc) + r'(?!\d)', base) is not None
        if casa_num or casa_doc:
            nm = (e.get('nome_razao_social') or '').strip()
            return ('#%s %s' % (num, nm)).strip() if num else (nm or '#? empresa')
    return None


def _ent_humano(n):
    n = float(n or 0)
    for u in ('B', 'KB', 'MB', 'GB'):
        if n < 1024:
            return ('%d %s' % (n, u)) if u == 'B' else ('%.1f %s' % (n, u))
        n /= 1024
    return '%.1f TB' % n


def _ent_parse_dt(s):
    if not s:
        return None
    s = str(s).strip().replace('Z', '+00:00')
    for cand in (s, s[:19]):
        try:
            dt = datetime.fromisoformat(cand)
            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt
        except ValueError:
            continue
    return None


def _ent_idade(s):
    dt = _ent_parse_dt(s)
    if not dt:
        return '—'
    delta = datetime.utcnow() - dt
    if delta.days >= 1:
        return 'há %d dia%s' % (delta.days, '' if delta.days == 1 else 's')
    h = delta.seconds // 3600
    if h >= 1:
        return 'há %d hora%s' % (h, '' if h == 1 else 's')
    m = (delta.seconds % 3600) // 60
    return ('há %d min' % m) if m else 'agora'


def _ent_montar(arquivos, empresas):
    """Metadados prontos p/ a tela, mais antigo primeiro."""
    itens = []
    for i in arquivos:
        nome = i.get('name') or ''
        ext = nome.rsplit('.', 1)[-1].lower() if '.' in nome else ''
        tlabel, ticone = _ENTRADA_TIPOS.get(ext, ('Outro', 'fa-file'))
        itens.append({
            'nome': nome, 'ext': ext, 'tipo': tlabel, 'icone': ticone,
            'tamanho': _ent_humano(i.get('size', 0)),
            'idade': _ent_idade(i.get('modified')),
            'empresa': _ent_casa_empresa(nome, empresas),
            '_ord': _ent_parse_dt(i.get('modified')) or datetime.max,
        })
    itens.sort(key=lambda x: x['_ord'])          # mais antigo em cima
    return itens


@configuracoes.route('/caixa-entrada')
@admin_required
def caixa_entrada():
    """Lista o _ENTRADA do Dropbox (só metadados). Filtro por tipo + busca."""
    from utils import dropbox_sync
    svc = dropbox_sync._service
    f_tipo = request.args.get('tipo', '').strip().lower()
    busca = request.args.get('busca', '').strip().lower()

    erro, itens, resumo, mais_antigo = None, [], {}, None
    if not svc.is_configured():
        erro = 'Dropbox não configurado (DROPBOX_APP_KEY / DROPBOX_REFRESH_TOKEN).'
    else:
        try:
            crus = svc.list_folder(svc.pasta_cert_novo())     # {ROOT}/_ENTRADA
        except Exception as exc:
            crus = []
            erro = 'Falha ao listar a caixa de entrada: %s' % str(exc)[:140]
        if not erro:
            empresas = execute_query(
                "SELECT numero_cliente, cpf_cnpj, nome_razao_social FROM clientes",
                fetch=True) or []
            todos = _ent_montar([i for i in crus if i.get('is_file')], empresas)
            # resumo por tipo + o mais antigo (item 5) — SOBRE O TOTAL, não o filtro
            for it in todos:
                resumo[it['tipo']] = resumo.get(it['tipo'], 0) + 1
            mais_antigo = todos[0]['idade'] if todos else None
            itens = todos
            if f_tipo:
                itens = [it for it in itens if it['ext'] == f_tipo]
            if busca:
                itens = [it for it in itens if busca in it['nome'].lower()]

    return render_template('configuracoes/caixa_entrada.html',
                           itens=itens, resumo=resumo, total=sum(resumo.values()),
                           mais_antigo=mais_antigo, erro=erro, f_tipo=f_tipo, busca=busca,
                           tipos=_ENTRADA_TIPOS)


@configuracoes.route('/caixa-entrada/renomear', methods=['POST'])
@admin_required
def caixa_entrada_renomear():
    """Renomeia um arquivo DENTRO do _ENTRADA (move_file). Recalcula o match e
    devolve o resultado. NÃO grava o nome no log (pode conter senha do .pfx)."""
    from utils import dropbox_sync
    svc = dropbox_sync._service
    if not svc.is_configured():
        return jsonify(ok=False, msg='Dropbox não configurado.'), 400
    data = request.get_json(silent=True) or {}
    atual = (data.get('atual') or '').strip()
    novo = (data.get('novo') or '').strip()
    if not atual or not novo:
        return jsonify(ok=False, msg='Informe o nome atual e o novo.'), 400
    # nomes de arquivo simples — sem barra, sem traversal (fica na própria _ENTRADA)
    if any(c in novo for c in '/\\') or novo in ('.', '..') or any(c in atual for c in '/\\'):
        return jsonify(ok=False, msg='Nome inválido.'), 400
    if atual == novo:
        return jsonify(ok=False, msg='O nome é o mesmo.'), 400

    base = svc.pasta_cert_novo()
    de, para = base + '/' + atual, base + '/' + novo
    try:
        if svc._path_exists(para):
            return jsonify(ok=False, msg='Já existe um arquivo com esse nome na caixa.'), 409
        if not svc.move_file(de, para):
            return jsonify(ok=False, msg='Falha ao renomear no Dropbox.'), 502
    except dropbox_sync.DropboxAuthError:
        return jsonify(ok=False, msg='Credenciais Dropbox inválidas ou expiradas.'), 401
    except dropbox_sync.DropboxError as exc:
        return jsonify(ok=False, msg='Erro no Dropbox: %s' % str(exc)[:120]), 502

    empresas = execute_query(
        "SELECT numero_cliente, cpf_cnpj, nome_razao_social FROM clientes", fetch=True) or []
    emp = _ent_casa_empresa(novo, empresas)
    # AUDITORIA: ação + módulo, SEM o nome (nomes de .pfx podem trazer a senha).
    registrar('escrita.renomeou_caixa_entrada', 'configuracoes', tabela=None,
              depois={'resultado': 'casou empresa' if emp else 'sem empresa identificada'})
    return jsonify(ok=True, empresa=emp)


@configuracoes.route('/dropbox/espaco/atualizar', methods=['POST'])
@admin_required
def dropbox_espaco_atualizar():
    """Força a releitura do espaço do Dropbox, ignorando o TTL do cache.

    Continua SOMENTE LEITURA: o 'atualizar' é do cache local, não do Dropbox.
    """
    from utils.dropbox_space import get_space
    return jsonify(get_space(force=True)), 200


# ===========================================================================
# USUÁRIOS
# ===========================================================================

@configuracoes.route('/usuarios/')
@admin_required
def usuarios():
    rows = execute_query(
        """SELECT u.id, u.nome, u.login, u.email, u.tipo_usuario, u.situacao, u.cargo,
                  u.senha_pendente,
                  GROUP_CONCAT(pa.nome ORDER BY pa.nome SEPARATOR ', ') AS perfis
             FROM usuarios u
             LEFT JOIN usuario_perfis up ON up.usuario_id = u.id
             LEFT JOIN perfis_acesso pa  ON pa.id = up.perfil_id
            WHERE u.classe_conta = 'FUNCIONARIO'
            GROUP BY u.id
            ORDER BY u.nome""",
        fetch=True,
    ) or []
    return render_template('configuracoes/usuarios_lista.html', usuarios=rows,
                           links=cadastro_token.listar(limite=50),
                           pendentes=_cadastros_pendentes(),
                           perfis_disponiveis=_perfis_ativos(),
                           departamentos_disponiveis=_departamentos_ativos(),
                           tipos_usuario=TIPOS_USUARIO,
                           senha_links=_links_senha_por_usuario(),
                           colabore_por_usuario=colabore_chaves.estado_por_usuario(),
                           hoje_iso=date.today().isoformat(),
                           csrf_token=_qrobo_csrf_token())


# ===========================================================================
# Q-COLABORE Parte 2 — link de cadastro + fila de pendentes
#
# Extensão desta MESMA tela: quem administra usuários é quem convida gente
# nova e quem aprova. Módulo Configurações intocado no resto.
#
# CSRF: reusa o helper de sessão do portal do Q-Robô, como o painel do
# escrita_fiscal já faz — gerar/revogar link é ação sensível e o helper custa
# uma linha. (SESSION_COOKIE_SAMESITE='Lax' é só a primeira barreira.)
# ===========================================================================

VALIDADE_LINK_HORAS = 72

# Os mesmos tipos do ENUM de usuarios.tipo_usuario — o painel de aprovação
# oferece exatamente estes, e o servidor recusa qualquer outro valor.
TIPOS_USUARIO = ('ADMIN', 'GERENTE', 'CONTADOR', 'ASSISTENTE', 'ESTAGIARIO')


class AprovacaoErro(RuntimeError):
    """Recusa amigável de uma aprovação (login tomado, e-mail duplicado, etc.).

    Não é erro de servidor: sobe até a rota, vira mensagem no painel e a pendência
    continua na fila para o admin corrigir. Distinta de Exception genérica, que
    seria um 500.
    """


def _perfis_ativos():
    return execute_query(
        "SELECT id, nome FROM perfis_acesso WHERE situacao='ATIVO' ORDER BY nome",
        fetch=True) or []


def _departamentos_ativos():
    return execute_query(
        'SELECT id, nome FROM departamentos WHERE ativo = 1 ORDER BY nome',
        fetch=True) or []


# --- Formatação SÓ para exibição no painel de análise. O banco continua com o
# valor cru (dígitos / datas ISO); estes helpers vestem os dados para o admin ler.
def _fmt_data(v):
    try:
        return v.strftime('%d/%m/%Y') if v is not None else None
    except Exception:
        return str(v) if v else None


def _fmt_datahora(v):
    try:
        return v.strftime('%d/%m/%Y %H:%M') if v is not None else None
    except Exception:
        return str(v) if v else None


def _fmt_cpf(v):
    d = re.sub(r'\D', '', v or '')
    return f'{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}' if len(d) == 11 else (v or None)


def _fmt_cel(v):
    d = re.sub(r'\D', '', v or '')
    if len(d) == 11:                       # (XX) X XXXX-XXXX
        return f'({d[:2]}) {d[2]} {d[3:7]}-{d[7:]}'
    if len(d) == 10:                       # (XX) XXXX-XXXX
        return f'({d[:2]}) {d[2:6]}-{d[6:]}'
    return v or None


def _gerar_link_senha(uid, admin_id, cur=None):
    """Gera um link tipo='SENHA' para o usuário e devolve (url, prefixo).

    cadastro_token.gerar já revoga os links de senha pendentes do mesmo usuário
    antes de criar o novo — é o que faz o "Gerar novo link" invalidar o anterior.
    A URL completa nasce em url_claro (via url_builder), então a lista pode
    reoferecer 'Copiar' enquanto o link estiver pendente.

    ``cur``: gera dentro de uma transação já aberta (usado quando o bloqueio de
    acesso e o link precisam ser um único commit atômico).
    """
    token, _link_id = cadastro_token.gerar(
        'SENHA', admin_id, usuario_id=uid, validade_horas=VALIDADE_LINK_HORAS,
        url_builder=lambda t: url_for('senha.formulario', token=t, _external=True),
        cur=cur)
    url = url_for('senha.formulario', token=token, _external=True)
    return url, token[:8]


def _links_senha_por_usuario():
    """{usuario_id: url_claro} dos links de SENHA pendentes (não usados/revogados/
    vencidos). Alimenta o botão 'Copiar link de senha' da lista de usuários."""
    rows = execute_query(
        """SELECT usuario_id, url_claro FROM cadastro_link
            WHERE tipo = 'SENHA' AND usuario_id IS NOT NULL
              AND usado_em IS NULL AND revogado_em IS NULL AND expira_em > NOW()""",
        fetch=True) or []
    return {r['usuario_id']: r['url_claro'] for r in rows if r.get('url_claro')}


def _qrobo_csrf_token():
    from routes.qrobo import csrf_token
    return csrf_token()


def _qrobo_csrf_ok():
    from routes.qrobo import csrf_valido
    return csrf_valido()


def _cadastros_pendentes():
    """Candidaturas aguardando decisão. Nasce vazia — o formulário é a Parte 3."""
    return execute_query(
        """SELECT p.id, p.nome_completo, p.login_escolhido, p.nick_escolhido,
                  p.email_pessoal, p.email_corporativo, p.celular_corporativo,
                  p.modalidade_trabalho, p.ja_funcionario,
                  p.criado_em, p.status,
                  GROUP_CONCAT(d.nome ORDER BY d.nome SEPARATOR ', ') AS departamentos
             FROM cadastro_pendente p
             LEFT JOIN cadastro_pendente_departamentos pd ON pd.pendente_id = p.id
             LEFT JOIN departamentos d ON d.id = pd.departamento_id
            WHERE p.status = 'PENDENTE'
            GROUP BY p.id
            ORDER BY p.criado_em""",
        fetch=True,
    ) or []


@configuracoes.route('/usuarios/link/gerar', methods=['POST'])
@admin_required
def usuario_link_gerar():
    """Gera um link de cadastro de uso único e devolve a URL UMA vez.

    Responde JSON de propósito: a URL com o token não passa por flash/session
    nem pela barra de endereços — some quando o modal fecha. Depois disso só o
    prefixo existe, aqui e no banco.
    """
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400

    destinatario = (request.form.get('destinatario') or '').strip() or None
    try:
        # url_builder: o util não conhece url_for. Passamos o construtor para que
        # url_claro (a cópia reenviável) nasça no mesmo INSERT do hash. Vive só
        # enquanto o link é pendente — some no uso/revogação/expiração.
        token, link_id = cadastro_token.gerar(
            'CADASTRO', current_user.id,
            destinatario=destinatario, validade_horas=VALIDADE_LINK_HORAS,
            url_builder=lambda tok: url_for('cadastro.formulario', token=tok,
                                            _external=True))
    except ValueError as exc:
        return jsonify(ok=False, msg=str(exc)), 400
    except Exception:
        # NUNCA ecoar a exceção crua: ela pode carregar o SQL e o parâmetro.
        logger.exception('[qcolabore] falha ao gerar link de cadastro')
        return jsonify(ok=False, msg='Não foi possível gerar o link agora.'), 500

    # A rota pública agora EXISTE (Parte 3) — url_for no lugar do path montado
    # à mão, como estava combinado desde a Parte 2.
    url = url_for('cadastro.formulario', token=token, _external=True)
    # Log com o ID e o prefixo. O token inteiro não entra em log nenhum.
    logger.info('[qcolabore] link de cadastro id=%s prefixo=%s gerado por %s.',
                link_id, token[:8], current_user.id)
    return jsonify(ok=True, url=url, prefixo=token[:8],
                   validade_horas=VALIDADE_LINK_HORAS)


@configuracoes.route('/usuarios/link/<int:link_id>/revogar', methods=['POST'])
@admin_required
def usuario_link_revogar(link_id):
    """Mata um link ainda não usado."""
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400
    if cadastro_token.revogar(link_id, current_user.id):
        logger.info('[qcolabore] link id=%s revogado por %s.', link_id, current_user.id)
        return jsonify(ok=True)
    # rowcount 0: já usado, já revogado, ou inexistente — nenhum deles é erro
    # de servidor, e a tela precisa recarregar para mostrar o estado real.
    return jsonify(ok=False, msg='Este link já foi usado ou revogado.'), 409


# ---------------------------------------------------------------------------
# Q-COLABORE Parte 5 — analisar / aprovar / recusar candidatura
# ---------------------------------------------------------------------------
@configuracoes.route('/usuarios/pendente/<int:pid>')
@admin_required
def usuario_pendente_detalhe(pid):
    """Todos os dados da candidatura para o painel de análise (JSON).

    Inclui os dados BANCÁRIOS: eles aparecem SÓ aqui, sob @admin_required, e nunca
    entram em log — por isso a resposta não é logada e o except de mais abaixo (nas
    rotas de decisão) não ecoa payload.
    """
    cand = execute_query(
        "SELECT * FROM cadastro_pendente WHERE id = %s AND status = 'PENDENTE'",
        (pid,), fetch=True, fetch_one=True)
    if not cand:
        return jsonify(ok=False, msg='Candidatura não encontrada ou já decidida.'), 404

    banco = execute_query(
        """SELECT banco_codigo, banco_nome, agencia, conta, conta_tipo,
                  titular_nome, titular_cpf, pix_tipo, pix_chave
             FROM usuario_dados_bancarios WHERE pendente_id = %s""",
        (pid,), fetch=True, fetch_one=True)

    deps = [r['departamento_id'] for r in (execute_query(
        "SELECT departamento_id FROM cadastro_pendente_departamentos WHERE pendente_id = %s",
        (pid,), fetch=True) or [])]

    campos = ('id', 'nome_completo', 'cpf', 'data_nascimento', 'login_escolhido',
              'nick_escolhido', 'email_pessoal', 'email_corporativo',
              'celular_pessoal', 'celular_corporativo', 'cep', 'logradouro',
              'numero', 'complemento', 'bairro', 'cidade', 'estado', 'pais',
              'modalidade_trabalho', 'ja_funcionario', 'criado_em')
    c = {k: cand.get(k) for k in campos}
    # Apresentação em padrão brasileiro (feita AQUI, no servidor, não no JS):
    c['data_nascimento'] = _fmt_data(cand.get('data_nascimento'))
    c['criado_em'] = _fmt_datahora(cand.get('criado_em'))
    c['cpf'] = _fmt_cpf(cand.get('cpf'))
    c['celular_pessoal'] = _fmt_cel(cand.get('celular_pessoal'))
    c['celular_corporativo'] = _fmt_cel(cand.get('celular_corporativo'))
    # AUDITORIA (D2): leitura — abriu a candidatura para análise. Grava SÓ o id
    # (nada do corpo: os dados bancários NUNCA vão a log).
    registrar('leitura.abriu_cadastro_colabore', 'colabore',
              tabela='cadastro_pendente', registro_id=pid)
    return jsonify(ok=True, cand=c, banco=banco, departamentos_pedidos=deps)


def _aprovar_pendente(pid, tipo, perfis, deps, admin_id):
    """Cria o usuário a partir da candidatura, em UMA transação (tudo ou nada).

    Devolve o id do usuário criado. Qualquer AprovacaoErro faz a transação inteira
    reverter (transacao() dá rollback em exceção) — nem usuário, nem perfis, nem
    mudança de status sobrevivem a uma falha no meio.
    """
    cand = execute_query('SELECT * FROM cadastro_pendente WHERE id = %s',
                         (pid,), fetch=True, fetch_one=True)
    if not cand or cand['status'] != 'PENDENTE':
        raise AprovacaoErro('Esta candidatura já foi decidida.')

    login = (cand['login_escolhido'] or '').strip().lower()
    email = ((cand.get('email_corporativo') or cand.get('email_pessoal') or '')
             .strip().lower()) or None
    if not email:
        raise AprovacaoErro('O candidato não informou e-mail — não dá para criar o '
                            'acesso sem um. Peça que complete o cadastro.')
    nick = cand.get('nick_escolhido')
    cpf = cand.get('cpf') or None
    telefone = cand.get('celular_corporativo') or cand.get('celular_pessoal')
    dep_primario = deps[0] if deps else None

    with transacao() as cur:
        # a) o login ainda está livre? (pode ter sido tomado desde a candidatura)
        cur.execute('SELECT id FROM usuarios WHERE login = %s', (login,))
        if cur.fetchone():
            raise AprovacaoErro(f'O login "{login}" já foi tomado. Ajuste o login da '
                                'candidatura e aprove de novo.')
        # e-mail e CPF são UNIQUE em usuarios: barra aqui com mensagem decente em
        # vez de deixar estourar erro de driver no meio do INSERT.
        cur.execute('SELECT id FROM usuarios WHERE email = %s', (email,))
        if cur.fetchone():
            raise AprovacaoErro(f'Já existe um usuário com o e-mail {email}.')
        if cpf:
            cur.execute('SELECT id FROM usuarios WHERE cpf = %s', (cpf,))
            if cur.fetchone():
                raise AprovacaoErro('Já existe um usuário com este CPF.')

        # b) cria a conta SEM senha e com login bloqueado (senha_pendente=1). A
        # Parte 6 manda o link de senha e zera a flag.
        cur.execute(
            """INSERT INTO usuarios
                 (nome, nick, email, login, senha_hash, senha_pendente,
                  tipo_usuario, classe_conta, situacao, cpf, telefone, departamento_id)
               VALUES (%s, %s, %s, %s, NULL, 1, %s, 'FUNCIONARIO', 'ATIVO', %s, %s, %s)""",
            (cand['nome_completo'], nick, email, login, tipo, cpf, telefone, dep_primario))
        uid = cur.lastrowid

        # perfis escolhidos pelo admin
        for pf in perfis:
            cur.execute('INSERT IGNORE INTO usuario_perfis (usuario_id, perfil_id) '
                        'VALUES (%s, %s)', (uid, pf))

        # e) departamentos DEFINITIVOS: reescreve o conjunto na tabela da pendência,
        # que segue ligada ao usuário (usuario_id) — é onde o conjunto completo mora
        # (usuarios só tem um departamento_id, gravado como primário acima).
        cur.execute('DELETE FROM cadastro_pendente_departamentos WHERE pendente_id = %s',
                    (pid,))
        for d in deps:
            cur.execute('INSERT INTO cadastro_pendente_departamentos '
                        '(pendente_id, departamento_id) VALUES (%s, %s)', (pid, d))

        # c) move os bancários da pendência para o usuário recém-criado
        cur.execute('UPDATE usuario_dados_bancarios SET usuario_id = %s '
                    'WHERE pendente_id = %s', (uid, pid))

        # d) marca a pendência aprovada. O WHERE status='PENDENTE' + rowcount==1
        # fecha a corrida: dois admins aprovando ao mesmo tempo → um só vence.
        cur.execute(
            """UPDATE cadastro_pendente
                  SET status = 'APROVADO', decidido_por = %s, decidido_em = NOW(),
                      usuario_id = %s
                WHERE id = %s AND status = 'PENDENTE'""",
            (admin_id, uid, pid))
        if cur.rowcount != 1:
            raise AprovacaoErro('Esta candidatura já foi decidida por outra pessoa.')

    return uid


@configuracoes.route('/usuarios/pendente/<int:pid>/aprovar', methods=['POST'])
@admin_required
def usuario_pendente_aprovar(pid):
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400

    tipo = (request.form.get('tipo_usuario') or '').strip().upper()
    if tipo not in TIPOS_USUARIO:
        return jsonify(ok=False, msg='Selecione um tipo de conta válido.'), 400
    perfis = [int(p) for p in request.form.getlist('perfis') if p.isdigit()]
    deps = [int(d) for d in request.form.getlist('departamentos') if d.isdigit()]

    try:
        uid = _aprovar_pendente(pid, tipo, perfis, deps, current_user.id)
    except AprovacaoErro as exc:
        # Falha esperada: pendência segue na fila para o admin ajustar.
        return jsonify(ok=False, msg=str(exc)), 409
    except Exception:
        # NUNCA ecoar a exceção crua — a transação pode carregar dados bancários.
        logger.exception('[qcolabore] falha ao aprovar pendência %s', pid)
        return jsonify(ok=False, msg='Não foi possível aprovar agora. Tente novamente.'), 500

    # Link de senha (Parte 6): a conta nasce sem senha; este link é como a pessoa
    # a define. Geração é NÃO-fatal — se falhar, o usuário já existe e o admin
    # gera outro pela lista ("Novo link de senha"). NÃO desfaz a aprovação.
    senha_url, senha_prefixo = None, None
    try:
        senha_url, senha_prefixo = _gerar_link_senha(uid, current_user.id)
    except Exception:
        logger.exception('[qcolabore] usuário %s criado, mas falhou o link de senha', uid)

    logger.info('[qcolabore] pendência %s aprovada por %s → usuário %s.',
                pid, current_user.id, uid)
    # AUDITORIA (D2): candidatura recebida foi TRATADA (aprovada). Sem dado bancário.
    registrar('escrita.aprovou_cadastro_colabore', 'colabore',
              tabela='cadastro_pendente', registro_id=pid,
              depois={'usuario_id': uid, 'tipo_usuario': tipo})
    return jsonify(ok=True, usuario_id=uid,
                   senha_url=senha_url, senha_prefixo=senha_prefixo)


@configuracoes.route('/usuarios/pendente/<int:pid>/recusar', methods=['POST'])
@admin_required
def usuario_pendente_recusar(pid):
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400

    motivo = (request.form.get('motivo') or '').strip()[:255]
    if not motivo:
        return jsonify(ok=False, msg='Diga um motivo curto para a recusa.'), 400

    # Não cria nada. A linha fica no banco como histórico (status RECUSADO) — some
    # só da FILA, que filtra por PENDENTE. rowcount==1 confirma que era pendente.
    with transacao() as cur:
        cur.execute(
            """UPDATE cadastro_pendente
                  SET status = 'RECUSADO', decisao_motivo = %s,
                      decidido_por = %s, decidido_em = NOW()
                WHERE id = %s AND status = 'PENDENTE'""",
            (motivo, current_user.id, pid))
        if cur.rowcount != 1:
            return jsonify(ok=False, msg='Candidatura já decidida ou inexistente.'), 409

    logger.info('[qcolabore] pendência %s recusada por %s.', pid, current_user.id)
    # AUDITORIA (D2): candidatura recebida foi TRATADA (recusada).
    registrar('escrita.recusou_cadastro_colabore', 'colabore',
              tabela='cadastro_pendente', registro_id=pid, depois={'motivo': motivo})
    return jsonify(ok=True)


@configuracoes.route('/usuarios/<int:uid>/senha-link', methods=['POST'])
@admin_required
def usuario_senha_link(uid):
    """Gera link de definição/redefinição de senha. Cobre três casos:

    - senha_pendente=1 (conta recém-aprovada, sem senha): só (re)emite o link.
    - senha_pendente=0 + bloquear=0: REDEFINIÇÃO "esqueci a senha" — emite o link;
      a senha atual CONTINUA valendo até a pessoa definir a nova.
    - senha_pendente=0 + bloquear=1: REDEFINIÇÃO com corte — marca senha_pendente=1
      E emite o link no MESMO commit; a pessoa é deslogada na próxima requisição
      (load_user recusa senha_pendente=1) e só volta pelo link. "Senha vazou /
      desligamento".

    gerar() revoga o link de senha pendente anterior. NUNCA loga token/senha.
    """
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400
    u = execute_query('SELECT id, senha_pendente FROM usuarios WHERE id = %s',
                      (uid,), fetch=True, fetch_one=True)
    if not u:
        return jsonify(ok=False, msg='Usuário não encontrado.'), 404

    bloquear = request.form.get('bloquear') == '1'

    # Guard anti-tiro-no-pé: bloquear o PRÓPRIO acesso derrubaria a sua sessão
    # agora — e se o link não for copiado, ninguém reentra. Redefinir a própria
    # senha SEM bloqueio segue permitido.
    if bloquear and uid == current_user.id:
        return jsonify(ok=False, msg='Você não pode bloquear o seu próprio acesso. '
                       'Redefina sem bloquear, ou peça a outro administrador.'), 409

    try:
        if bloquear and u['senha_pendente'] == 0:
            # Bloqueio + link num ÚNICO commit: ou a pessoa perde o acesso E ganha
            # o caminho de volta, ou nada muda.
            with transacao() as cur:
                cur.execute('UPDATE usuarios SET senha_pendente = 1 WHERE id = %s',
                            (uid,))
                url, prefixo = _gerar_link_senha(uid, current_user.id, cur=cur)
        else:
            url, prefixo = _gerar_link_senha(uid, current_user.id)
    except Exception:
        logger.exception('[qcolabore] falha ao gerar link de senha do usuário %s', uid)
        return jsonify(ok=False, msg='Não foi possível gerar o link agora.'), 500

    logger.info('[qcolabore] link de senha para usuário %s por %s (bloquear=%s).',
                uid, current_user.id, int(bloquear))
    return jsonify(ok=True, url=url, prefixo=prefixo, bloqueado=bool(bloquear))


# ===========================================================================
# Q-COLABORE F2 — chave do agente POR FUNCIONÁRIO (gerar/regenerar/revogar)
#
# A chave autentica o agente instalado na máquina da pessoa (POST /api/colabore/
# enviar). O segredo aparece UMA vez aqui, na resposta; depois só o prefixo. Toda
# a lógica de hash+prefixo mora em utils/colabore_chaves. Admin, trava no servidor.
# ===========================================================================
@configuracoes.route('/usuarios/<int:uid>/colabore/gerar', methods=['POST'])
@admin_required
def usuario_colabore_gerar(uid):
    """Gera (ou regenera) a chave do Q-Colabore do funcionário. Devolve o segredo
    UMA vez. ``regerar=1`` confirma a rotação (invalida a chave anterior). A data
    de corte (data_inicio_captura) vem do form; padrão = hoje no servidor."""
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400

    regerar = request.form.get('regerar') == '1'
    data_inicio = (request.form.get('data_inicio') or '').strip() or None
    ip, ua = colabore_chaves.contexto_request()

    res = colabore_chaves.gerar_chave(uid, current_user.id, regerar=regerar,
                                      data_inicio=data_inicio, ip=ip, user_agent=ua)
    if not res.get('ok'):
        erro = res.get('erro')
        if erro == 'ja_existe':
            # A tela precisa avisar e pedir a 2ª confirmação antes de rotacionar.
            return jsonify(ok=False, erro='ja_existe',
                           msg='Este funcionário já tem uma chave. Gerar nova '
                               'INVALIDA a atual — confirme para continuar.'), 409
        mapa = {'usuario_inexistente': ('Usuário não encontrado.', 404),
                'data_invalida': ('Data de corte inválida.', 400),
                'data_futura': ('A data de corte não pode ser no futuro.', 400)}
        msg, cod = mapa.get(erro, ('Não foi possível gerar a chave agora.', 500))
        return jsonify(ok=False, erro=erro, msg=msg), cod

    di = res.get('data_inicio_captura')
    # AUDITORIA: ação + funcionário + versão + data de corte. NUNCA o segredo
    # (nem parcial além do prefixo). CAMPOS_SENSIVEIS já mascara 'token'/'chave'.
    registrar('escrita.gerou_chave_colabore', 'colabore',
              tabela='colabore_config', registro_id=uid,
              depois={'acao': res['acao'], 'versao': res['versao'],
                      'prefixo': res['prefixo'],
                      'data_inicio_captura': di.isoformat() if hasattr(di, 'isoformat') else di})
    return jsonify(ok=True, token=res['token'], prefixo=res['prefixo'],
                   versao=res['versao'], acao=res['acao'],
                   data_inicio_captura=di.isoformat() if hasattr(di, 'isoformat') else di)


@configuracoes.route('/usuarios/<int:uid>/colabore/revogar', methods=['POST'])
@admin_required
def usuario_colabore_revogar(uid):
    """Revoga a chave do funcionário (ativo=0). O agente passa a receber 403. O
    prefixo continua visível (histórico); gerar nova reativa e rotaciona."""
    if not _qrobo_csrf_ok():
        return jsonify(ok=False, msg='Formulário expirado. Recarregue a página.'), 400

    res = colabore_chaves.revogar_chave(uid)
    if not res.get('ok'):
        return jsonify(ok=False, msg='Não foi possível revogar agora.'), 500
    registrar('escrita.revogou_chave_colabore', 'colabore',
              tabela='colabore_config', registro_id=uid, depois={'ativo': 0})
    return jsonify(ok=True)


@configuracoes.route('/usuarios/novo', methods=['GET', 'POST'])
@admin_required
def usuario_novo():
    perfis = execute_query(
        "SELECT id, nome FROM perfis_acesso WHERE situacao='ATIVO' ORDER BY nome",
        fetch=True,
    ) or []
    clientes = execute_query(
        "SELECT id, nome_razao_social FROM clientes WHERE situacao='ATIVO' ORDER BY nome_razao_social",
        fetch=True,
    ) or []

    if request.method == 'POST':
        nome     = request.form.get('nome', '').strip()
        login    = request.form.get('login', '').strip().lower().replace(' ', '')
        email    = request.form.get('email', '').strip().lower()
        senha    = request.form.get('senha', '').strip()
        tipo     = request.form.get('tipo_usuario', 'ASSISTENTE')
        cargo    = request.form.get('cargo', '').strip()
        telefone = request.form.get('telefone', '').strip()
        perfis_sel = request.form.getlist('perfis')
        empresas_sel = request.form.getlist('empresas')

        if not nome or not login or not email or not senha:
            flash('Nome, login, e-mail e senha são obrigatórios.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=None,
                                   perfis_usuario=set(), empresas_usuario=set())

        # Verifica login duplicado
        existe_login = execute_query(
            "SELECT id FROM usuarios WHERE login = %s", (login,), fetch=True, fetch_one=True
        )
        if existe_login:
            flash('Já existe um usuário com este login.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=None,
                                   perfis_usuario=set(), empresas_usuario=set())

        # Verifica e-mail duplicado
        existe = execute_query(
            "SELECT id FROM usuarios WHERE email = %s", (email,), fetch=True, fetch_one=True
        )
        if existe:
            flash('Já existe um usuário com este e-mail.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=None,
                                   perfis_usuario=set(), empresas_usuario=set())

        uid = execute_query(
            """INSERT INTO usuarios (nome, login, email, senha_hash, tipo_usuario, situacao, cargo, telefone)
               VALUES (%s, %s, %s, %s, %s, 'ATIVO', %s, %s)""",
            (nome, login, email, hash_password(senha), tipo, cargo, telefone),
        )

        if not uid:
            flash('Erro ao criar o usuário. Verifique os dados e tente novamente.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=None,
                                   perfis_usuario=set(), empresas_usuario=set())

        # Perfis
        for pid in perfis_sel:
            execute_query(
                "INSERT IGNORE INTO usuario_perfis (usuario_id, perfil_id) VALUES (%s, %s)",
                (uid, pid), fetch=False,
            )

        # Empresas permitidas
        for cid in empresas_sel:
            execute_query(
                "INSERT IGNORE INTO usuario_empresas_permitidas (usuario_id, cliente_id) VALUES (%s, %s)",
                (uid, cid), fetch=False,
            )

        flash(f'Usuário "{nome}" criado com sucesso.', 'success')
        return redirect(url_for('configuracoes.usuarios'))

    return render_template('configuracoes/usuario_form.html',
                           perfis=perfis, clientes=clientes, usuario=None,
                           perfis_usuario=set(), empresas_usuario=set())


@configuracoes.route('/usuarios/<int:uid>/editar', methods=['GET', 'POST'])
@admin_required
def usuario_editar(uid):
    usuario = execute_query(
        "SELECT * FROM usuarios WHERE id = %s", (uid,), fetch=True, fetch_one=True
    )
    if not usuario:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('configuracoes.usuarios'))

    perfis = execute_query(
        "SELECT id, nome FROM perfis_acesso WHERE situacao='ATIVO' ORDER BY nome",
        fetch=True,
    ) or []
    clientes = execute_query(
        "SELECT id, nome_razao_social FROM clientes WHERE situacao='ATIVO' ORDER BY nome_razao_social",
        fetch=True,
    ) or []
    perfis_usuario = {
        r['perfil_id']
        for r in (execute_query(
            "SELECT perfil_id FROM usuario_perfis WHERE usuario_id = %s", (uid,), fetch=True,
        ) or [])
    }
    empresas_usuario = {
        r['cliente_id']
        for r in (execute_query(
            "SELECT cliente_id FROM usuario_empresas_permitidas WHERE usuario_id = %s",
            (uid,), fetch=True,
        ) or [])
    }

    if request.method == 'POST':
        nome     = request.form.get('nome', '').strip()
        login    = request.form.get('login', '').strip().lower().replace(' ', '')
        email    = request.form.get('email', '').strip().lower()
        tipo     = request.form.get('tipo_usuario', 'ASSISTENTE')
        cargo    = request.form.get('cargo', '').strip()
        telefone = request.form.get('telefone', '').strip()
        situacao = request.form.get('situacao', 'ATIVO')
        nova_senha = request.form.get('nova_senha', '').strip()
        perfis_sel   = request.form.getlist('perfis')
        empresas_sel = request.form.getlist('empresas')

        if not nome or not login or not email:
            flash('Nome, login e e-mail são obrigatórios.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=usuario,
                                   perfis_usuario=perfis_usuario,
                                   empresas_usuario=empresas_usuario)

        # Verifica login duplicado (excluindo o próprio)
        existe_login = execute_query(
            "SELECT id FROM usuarios WHERE login = %s AND id != %s",
            (login, uid), fetch=True, fetch_one=True,
        )
        if existe_login:
            flash('Já existe outro usuário com este login.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=usuario,
                                   perfis_usuario=perfis_usuario,
                                   empresas_usuario=empresas_usuario)

        # Verifica e-mail duplicado (excluindo o próprio)
        existe = execute_query(
            "SELECT id FROM usuarios WHERE email = %s AND id != %s",
            (email, uid), fetch=True, fetch_one=True,
        )
        if existe:
            flash('Já existe outro usuário com este e-mail.', 'danger')
            return render_template('configuracoes/usuario_form.html',
                                   perfis=perfis, clientes=clientes, usuario=usuario,
                                   perfis_usuario=perfis_usuario,
                                   empresas_usuario=empresas_usuario)

        if nova_senha:
            execute_query(
                """UPDATE usuarios SET nome=%s, login=%s, email=%s, senha_hash=%s, tipo_usuario=%s,
                                       situacao=%s, cargo=%s, telefone=%s
                    WHERE id=%s""",
                (nome, login, email, hash_password(nova_senha), tipo, situacao, cargo, telefone, uid),
                fetch=False,
            )
        else:
            execute_query(
                """UPDATE usuarios SET nome=%s, login=%s, email=%s, tipo_usuario=%s,
                                       situacao=%s, cargo=%s, telefone=%s
                    WHERE id=%s""",
                (nome, login, email, tipo, situacao, cargo, telefone, uid),
                fetch=False,
            )

        # Atualiza perfis
        execute_query("DELETE FROM usuario_perfis WHERE usuario_id = %s", (uid,), fetch=False)
        for pid in perfis_sel:
            execute_query(
                "INSERT IGNORE INTO usuario_perfis (usuario_id, perfil_id) VALUES (%s, %s)",
                (uid, pid), fetch=False,
            )

        # Atualiza empresas
        execute_query(
            "DELETE FROM usuario_empresas_permitidas WHERE usuario_id = %s", (uid,), fetch=False,
        )
        for cid in empresas_sel:
            execute_query(
                "INSERT IGNORE INTO usuario_empresas_permitidas (usuario_id, cliente_id) VALUES (%s, %s)",
                (uid, cid), fetch=False,
            )

        flash(f'Usuário "{nome}" atualizado com sucesso.', 'success')
        return redirect(url_for('configuracoes.usuarios'))

    return render_template('configuracoes/usuario_form.html',
                           perfis=perfis, clientes=clientes, usuario=usuario,
                           perfis_usuario=perfis_usuario,
                           empresas_usuario=empresas_usuario)


@configuracoes.route('/usuarios/<int:uid>/toggle', methods=['POST'])
@admin_required
def usuario_toggle(uid):
    """Ativa/desativa um usuário."""
    u = execute_query("SELECT situacao FROM usuarios WHERE id=%s", (uid,), fetch=True, fetch_one=True)
    if not u:
        return jsonify(ok=False, msg='Usuário não encontrado.'), 404
    nova = 'INATIVO' if u['situacao'] == 'ATIVO' else 'ATIVO'
    execute_query("UPDATE usuarios SET situacao=%s WHERE id=%s", (nova, uid), fetch=False)
    return jsonify(ok=True, situacao=nova)


# ===========================================================================
# PERFIS DE ACESSO
# ===========================================================================

@configuracoes.route('/perfis/')
@admin_required
def perfis():
    rows = execute_query(
        """SELECT pa.id, pa.nome, pa.descricao, pa.situacao,
                  COUNT(DISTINCT up.usuario_id) AS qtd_usuarios,
                  COUNT(DISTINCT pp.permissao_codigo) AS qtd_permissoes
             FROM perfis_acesso pa
             LEFT JOIN usuario_perfis up ON up.perfil_id = pa.id
             LEFT JOIN perfil_permissoes pp ON pp.perfil_id = pa.id
            GROUP BY pa.id
            ORDER BY pa.nome""",
        fetch=True,
    ) or []
    return render_template('configuracoes/perfis_lista.html', perfis=rows)


@configuracoes.route('/perfis/novo', methods=['GET', 'POST'])
@admin_required
def perfil_novo():
    if request.method == 'POST':
        nome      = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        perms_sel = request.form.getlist('permissoes')

        if not nome:
            flash('O nome do perfil é obrigatório.', 'danger')
            return render_template('configuracoes/perfil_form.html',
                                   catalog=PERMISSION_CATALOG, perfil=None, perms_perfil=set())

        pid = execute_query(
            "INSERT INTO perfis_acesso (nome, descricao) VALUES (%s, %s)",
            (nome, descricao),
        )
        for cod in perms_sel:
            execute_query(
                "INSERT IGNORE INTO perfil_permissoes (perfil_id, permissao_codigo) VALUES (%s, %s)",
                (pid, cod), fetch=False,
            )

        flash(f'Perfil "{nome}" criado com sucesso.', 'success')
        return redirect(url_for('configuracoes.perfis'))

    return render_template('configuracoes/perfil_form.html',
                           catalog=PERMISSION_CATALOG, perfil=None, perms_perfil=set())


@configuracoes.route('/perfis/<int:pid>/editar', methods=['GET', 'POST'])
@admin_required
def perfil_editar(pid):
    perfil = execute_query(
        "SELECT * FROM perfis_acesso WHERE id = %s", (pid,), fetch=True, fetch_one=True,
    )
    if not perfil:
        flash('Perfil não encontrado.', 'danger')
        return redirect(url_for('configuracoes.perfis'))

    perms_perfil = {
        r['permissao_codigo']
        for r in (execute_query(
            "SELECT permissao_codigo FROM perfil_permissoes WHERE perfil_id = %s",
            (pid,), fetch=True,
        ) or [])
    }

    if request.method == 'POST':
        nome      = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        situacao  = request.form.get('situacao', 'ATIVO')
        perms_sel = request.form.getlist('permissoes')

        if not nome:
            flash('O nome do perfil é obrigatório.', 'danger')
            return render_template('configuracoes/perfil_form.html',
                                   catalog=PERMISSION_CATALOG, perfil=perfil,
                                   perms_perfil=perms_perfil)

        execute_query(
            "UPDATE perfis_acesso SET nome=%s, descricao=%s, situacao=%s WHERE id=%s",
            (nome, descricao, situacao, pid), fetch=False,
        )
        execute_query(
            "DELETE FROM perfil_permissoes WHERE perfil_id = %s", (pid,), fetch=False,
        )
        for cod in perms_sel:
            execute_query(
                "INSERT IGNORE INTO perfil_permissoes (perfil_id, permissao_codigo) VALUES (%s, %s)",
                (pid, cod), fetch=False,
            )

        flash(f'Perfil "{nome}" atualizado com sucesso.', 'success')
        return redirect(url_for('configuracoes.perfis'))

    return render_template('configuracoes/perfil_form.html',
                           catalog=PERMISSION_CATALOG, perfil=perfil, perms_perfil=perms_perfil)


@configuracoes.route('/perfis/<int:pid>/excluir', methods=['POST'])
@admin_required
def perfil_excluir(pid):
    p = execute_query("SELECT nome FROM perfis_acesso WHERE id=%s", (pid,), fetch=True, fetch_one=True)
    if not p:
        flash('Perfil não encontrado.', 'danger')
        return redirect(url_for('configuracoes.perfis'))
    execute_query("DELETE FROM perfis_acesso WHERE id=%s", (pid,), fetch=False)
    flash(f'Perfil "{p["nome"]}" excluído.', 'success')
    return redirect(url_for('configuracoes.perfis'))
